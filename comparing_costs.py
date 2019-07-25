'''
Programme to compare a specific year or total wlc information. It produces a wb output with data and calculations
only i.e. no graph. The output shows changes to wlc in relation 1) overall figures, 2) change between quarters,
3) percent change are highlighted in red if change is greater/less than £100m/-£100m or percentage change greater/less
than 5%/-5% of project value

It is from the data placed into the output document that a simple bard chart can be built to show the most significant
changes in cost since the previous quarter.
'''

from bcompiler.utils import project_data_from_master
from openpyxl import Workbook
from openpyxl.styles import Font


''' One of key functions used for calculating which quarter to baseline data from...
Function returns a dictionary structured in the following way project name[('latest quarter info', 'latest bc'), 
('last quarter info', 'last bc'), ('last baseline quarter info', 'last baseline bc'), ('oldest quarter info', 
'oldest bc')] depending on the amount information available in the data. Only the first three key values are returned, 
to ensure consistency (which is helpful later).'''
def bc_ref_stages(proj_list, q_masters_dict_list):

    output_dict = {}

    for name in proj_list:
        #print(name)
        all_list = []      # format [('quarter info': 'bc')] across all masters including project
        bl_list = []        # format ['bc', 'bc'] across all masters. bl_list_2 removes duplicates
        ref_list = []       # format as for all list but only contains the three tuples of interest
        for master in q_masters_dict_list:
            try:
                bc_stage = master[name]['BICC approval point']
                quarter = master[name]['Reporting period (GMPP - Snapshot Date)']
                tuple = (quarter, bc_stage)
                all_list.append(tuple)
            except KeyError:
                pass

        for i in range(0, len(all_list)):
            bl_list.append(all_list[i][1])

        '''below lines of text from stackoverflow. Question, remove duplicates in python list while 
        preserving order'''
        seen = set()
        seen_add = seen.add
        bl_list_2 = [x for x in bl_list if not (x in seen or seen_add(x))]

        ref_list.insert(0, all_list[0])     # puts the latest info into the list first

        try:
            ref_list.insert(1, all_list[1])    # puts that last info into the list
        except IndexError:
            ref_list.insert(1, all_list[0])

        if len(bl_list_2) == 1:                     # puts oldest info into list (as basline if no baseline)
            ref_list.insert(2, all_list[-1])
        else:
            for i in range(0, len(all_list)):      # puts in baseline
                if all_list[i][1] == bl_list[0]:
                    ref_list.insert(2, all_list[i])

        '''there is a hack here i.e. returning only first three in ref_list. There's a bug which I don't fully 
        understand, but this solution is hopefully good enough for now'''
        output_dict[name] = ref_list[0:3]

    return output_dict

'''Another key function used for calcualting which quarter to baseline data from...
Fuction returns a dictionay structured in the following way project_name[n,n,n]. The n (number) values denote where 
the relevant quarter master dictionary is positions in the list of master dictionaries'''
def get_master_baseline_dict(proj_list, q_masters_dict_list, baseline_dict_list):
    output_dict = {}

    for name in proj_list:
        master_q_list = []
        for key in baseline_dict_list[name]:
            for x, master in enumerate(q_masters_dict_list):
                try:
                    quarter = master[name]['Reporting period (GMPP - Snapshot Date)']
                    if quarter == key[0]:
                        master_q_list.append(x)
                except KeyError:
                    pass

        output_dict[name] = master_q_list

    return output_dict


def compare(data_1, data_2):
    wb = Workbook()
    ws = wb.active

    for i, name in enumerate(data_1):
        '''place project names into ws'''
        ws.cell(row=i+2, column=1).value = name

        '''loop for placing wlc data into ws. highlight changes between quarters in red'''
        latest_wlc = data_1[name]
        try:
            last_wlc = data_2[name]
        except KeyError:
            last_wlc = 'None'

        ws.cell(row=i + 2, column=2).value = latest_wlc

        if latest_wlc != last_wlc:
            ws.cell(row=i + 2, column=2).font = red_text

        if name in data_2.keys():
            try:
                ws.cell(row=i + 2, column=3).value = last_wlc
                change = latest_wlc - last_wlc
                if last_wlc > 0:
                    percent_change = (latest_wlc - last_wlc)/last_wlc
                else:
                    percent_change = (latest_wlc - last_wlc)/(last_wlc + 1)
                ws.cell(row=i + 2, column=4).value = change
                ws.cell(row=i + 2, column=5).value = percent_change
                if change >= 100 or change <= -100:
                    ws.cell(row=i + 2, column=4).font = red_text
                if percent_change >= 0.05 or percent_change <= -0.05:
                    ws.cell(row=i + 2, column=5).font = red_text
            except TypeError:
                pass
        else:
            ws.cell(row=i + 2, column=3).value = last_wlc

    ws.cell(row=1, column=1).value = 'Project Name'
    ws.cell(row=1, column=2).value = 'Latest Quarter'
    ws.cell(row=1, column=3).value = 'Baseline Quarter'
    ws.cell(row=1, column=4).value = 'Change'
    ws.cell(row=1, column=5).value = 'Percentage Change'
    return wb

def get_yearly_costs(proj_list, q_masters_dict_list, cost_list, year, index):
    output_dict = {}
    for proj_name in proj_list:
        project_dict = q_masters_dict_list[q_masters_list[proj_name][index]][proj_name]
        total = 0
        for type in cost_list:
            if year + type in project_dict.keys():
                cost = project_dict[year + type]
                try:
                    total = total + cost
                except TypeError:
                    pass

        output_dict[proj_name] = total

    return output_dict

def get_wlc(proj_list, q_masters_dict_list, key, index):
    output_dict = {}
    for proj_name in proj_list:
        proj_dict = q_masters_dict_list[q_masters_list[proj_name][index]][proj_name]
        total = proj_dict[key]
        output_dict[proj_name] = total

    return output_dict

red_text = Font(color="FF0000")

'''INSTRUCTIONS FOR RUNNING PROGRAMME'''

'''1) load all master quarter data files here'''
q1_1920 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2019_wip_'
                                   '(18_7_19).xlsx')
q4_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2018.xlsx')
q3_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2018.xlsx')
q2_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_2_2018.xlsx')
q1_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2018.xlsx')
q4_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2017.xlsx')
q3_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2017.xlsx')
q2_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_2_2017.xlsx')
q1_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2017.xlsx')
q4_1617 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2016.xlsx')
q3_1617 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2016.xlsx')

'''2) Include in the below list, as the variable names, those quarters to include in analysis
NOTE - the comparison of financial totals us the 'bespoke' list consistent cost reporting (in nominal figures) was
 only achieved in q1_1819, so it would be incorrect to compare figures beyond this'''
list_of_dicts_all = [q1_1920 ,q4_1819, q3_1819, q2_1819, q1_1819, q4_1718, q3_1718, q2_1718, q1_1718, q4_1617, q3_1617]
list_of_dicts_bespoke = [q1_1920, q4_1819, q3_1819, q2_1819, q1_1819]

# '''1) specify file paths to where master data for analysis is stored.'''
# latest_q_data = project_data_from_master("C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2019"
#                                          "_wip_(18_7_19).xlsx")
# other_q_data = project_data_from_master("C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2018.xlsx")

proj_names_all = list(q1_1920.keys())

baseline_bc = bc_ref_stages(proj_names_all, list_of_dicts_bespoke)
q_masters_list = get_master_baseline_dict(proj_names_all, list_of_dicts_bespoke, baseline_bc)

'''2) decide which output you require'''

'''in year cost lists is chosen through the cost list. No not change.'''
cost_list = [' RDEL Forecast Total', ' CDEL Forecast Total', ' Forecast Non-Gov']

'''OPTION ONE - for comparing in year costs'''
'''in year income list is chosen through the income list. No not change.'''
income_list = [' Forecast - Income both Revenue and Capital']

'''chose financial year of interest. change accordingly. needs to be in format of YY-YY'''
year_interest = '23-24'

'''get fy information by entering the appropriate variables'''
latest_fy = get_yearly_costs(proj_names_all, list_of_dicts_bespoke, cost_list, year_interest, 0)
baseline_fy = get_yearly_costs(proj_names_all, list_of_dicts_bespoke, cost_list, year_interest, 2)

'''OPTION TWO - for wlc costs'''

'''chose wlc cost key of interest from master data. Get information by entering appropriate variables below'''
wlc_key = 'Total Forecast'
latest_wlc = get_wlc(proj_names_all, list_of_dicts_bespoke, wlc_key, 0)
baseline_wlc = get_wlc(proj_names_all, list_of_dicts_bespoke, wlc_key, 2)

'''3) enter desired variables into the compare function i.e. enter either one_fy, two_fy or one_wlc, two_wlc and 
specify file path for where output document to be saved'''
output = compare(latest_fy, baseline_fy)

output.save("C:\\Users\\Standalone\\general\\Q1_1920_fy_23_24_comparison_against_baseline.xlsx")