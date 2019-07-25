'''
Programme that creates financial cost profile for individual projects. Follow instructions below.

Recently modified so that the three different cost profiles calculated are i) latest, ii) last, iii) baseline.
'''

from openpyxl import Workbook
from bcompiler.utils import project_data_from_master
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font

''' One of key functions used for calculating which quarter to baseline data from...
Function returns a dictionary structured in the following way project name[('latest quarter info', 'latest bc'), 
('last quarter info', 'last bc'), ('last baseline quarter info', 'last baseline bc'), ('oldest quarter info', 
'oldest bc')] depending on the amount information available in the data. Only the first three key values are returned, 
to ensure consistency (which is helpful later).'''
def bc_ref_stages(proj_list, q_masters_dict_list):

    output_dict = {}

    for name in proj_list:
        #print(name)
        all_list = []      # format [('quarter info': 'bc')] across all masters including project
        bl_list = []        # format ['bc', 'bc'] across all masters. bl_list_2 removes duplicates
        ref_list = []       # format as for all list but only contains the three tuples of interest
        for master in q_masters_dict_list:
            try:
                bc_stage = master[name]['BICC approval point']
                quarter = master[name]['Reporting period (GMPP - Snapshot Date)']
                tuple = (quarter, bc_stage)
                all_list.append(tuple)
            except KeyError:
                pass

        for i in range(0, len(all_list)):
            bl_list.append(all_list[i][1])

        '''below lines of text from stackoverflow. Question, remove duplicates in python list while 
        preserving order'''
        seen = set()
        seen_add = seen.add
        bl_list_2 = [x for x in bl_list if not (x in seen or seen_add(x))]

        ref_list.insert(0, all_list[0])     # puts the latest info into the list first

        try:
            ref_list.insert(1, all_list[1])    # puts that last info into the list
        except IndexError:
            ref_list.insert(1, all_list[0])

        if len(bl_list_2) == 1:                     # puts oldest info into list (as basline if no baseline)
            ref_list.insert(2, all_list[-1])
        else:
            for i in range(0, len(all_list)):      # puts in baseline
                if all_list[i][1] == bl_list[0]:
                    ref_list.insert(2, all_list[i])

        '''there is a hack here i.e. returning only first three in ref_list. There's a bug which I don't fully 
        understand, but this solution is hopefully good enough for now'''
        output_dict[name] = ref_list[0:3]

    return output_dict

'''Another key function used for calcualting which quarter to baseline data from...
Fuction returns a dictionay structured in the following way project_name[n,n,n]. The n (number) values denote where 
the relevant quarter master dictionary is positions in the list of master dictionaries'''
def get_master_baseline_dict(proj_list, q_masters_dict_list, baseline_dict_list):
    output_dict = {}

    for name in proj_list:
        master_q_list = []
        for key in baseline_dict_list[name]:
            for x, master in enumerate(q_masters_dict_list):
                try:
                    quarter = master[name]['Reporting period (GMPP - Snapshot Date)']
                    if quarter == key[0]:
                        master_q_list.append(x)
                except KeyError:
                    pass

        output_dict[name] = master_q_list

    return output_dict


def financial_dict(proj_list, q_masters_dict_list, q_masters_list, cells_to_capture, index):
    '''
    the creation of a mini dictionary containing financial information. This is done via two functions; this one
    and the one title financial_info ( directly below ).
    :param name_list: list of project names
    :param master_data: master data for quarter of interest
    :param cells_to_capture: financial info key names. see lists below
    :return:
    '''

    output_dict = {}
    for proj_name in proj_list:
        master_data = q_masters_dict_list[q_masters_list[proj_name][index]]
        get_dict_info = financial_info(proj_name, master_data, cells_to_capture)
        output_dict[proj_name] = get_dict_info

    return output_dict

def financial_info(name, master_data, cells_to_capture):
    '''
    function that creates dictionary containing financial {key : value} information.
    names = name of project
    master_data = master data set
    cells_to_capture = lists of keys of interest
    '''

    output_dict = {}

    if name in master_data.keys():
        for item in master_data[name]:
            if item in cells_to_capture:
                if master_data[name][item] is None:
                    output_dict[item] = 0
                else:
                    value = master_data[name][item]
                    output_dict[item] = value

    else:
        for item in cells_to_capture:
            output_dict[item] = 0

    return output_dict

def calculate_totals(name, fin_data):
    '''
    :param name: project name
    :param fin_data: mini project financial dictionary
    :return: a total number for rdel, cdel, and ng spend each financial year
    '''

    working_data = fin_data[name]
    rdel_list = []
    cdel_list = []
    ng_list = []

    for rdel in capture_rdel:
        try:
            rdel_list.append(working_data[rdel])
        except KeyError:
            rdel_list.append(int(0))
    for cdel in capture_cdel:
        try:
            cdel_list.append(working_data[cdel])
        except KeyError:
            cdel_list.append(int(0))
    for ng in capture_ng:
        try:
            ng_list.append(working_data[ng])
        except KeyError:
            ng_list.append(int(0))

    total_list = []
    for i in range(len(rdel_list)):
        total = rdel_list[i] + cdel_list[i] + ng_list[i]
        total_list.append(total)

    return total_list

def calculate_income_totals(name, fin_data):
    '''
    :param name: project name
    :param fin_data: mini project financial dictionary
    :return: a total number for income spending each financial year
    '''

    working_data = fin_data[name]
    income_list = []

    for income in capture_income:
        try:
            income_list.append(working_data[income])
        except KeyError:
            income_list.append(int(0))

    return income_list

def place_in_excel(name, latest_fin_data, last_fin_data, baseline_fin_data):
    '''
    function places all data into excel spreadsheet and creates chart.
    data is placed into sheet in reverse order (see how data_list is ordered) so that most recent
    data is displayed on right hand side of the data table
    '''

    wb = Workbook()
    ws = wb.active
    data_list = [baseline_fin_data, last_fin_data, latest_fin_data]
    count = 0

    '''places in raw/reported data'''
    for data in data_list:
        for i, key in enumerate(capture_rdel):
            try:
                ws.cell(row=i+3, column=2+count, value=data[name][key])
            except KeyError:
                ws.cell(row=i + 3, column=2 + count, value=0)
        for i, key in enumerate(capture_cdel):
            try:
                ws.cell(row=i+3, column=3+count, value=data[name][key])
            except KeyError:
                ws.cell(row=i + 3, column=3 + count, value=0)
        for i, key in enumerate(capture_ng):
            try:
                ws.cell(row=i+3, column=4+count, value=data[name][key])
            except KeyError:
                ws.cell(row=i + 3, column=4 + count, value=0)
        count += 4

    '''places in totals'''
    baseline_totals = calculate_totals(name, baseline_fin_data)
    last_q_totals = calculate_totals(name, last_fin_data)
    latest_q_totals = calculate_totals(name, latest_fin_data)

    total_list = [baseline_totals, last_q_totals, latest_q_totals]

    c = 0
    for l in total_list:
        for i, total in enumerate(l):
            ws.cell(row=i+3, column=5+c, value=total)
        c += 4

    '''labeling data in table'''

    labeling_list_quarter = ['Baseline', 'Last Quarter', 'Latest quarter']

    ws.cell(row=1, column=2, value=labeling_list_quarter[0])
    ws.cell(row=1, column=6, value=labeling_list_quarter[1])
    ws.cell(row=1, column=10, value=labeling_list_quarter[2])

    labeling_list_type = ['RDEL', 'CDEL', 'Non-Gov', 'Total']
    repeat = 3
    c = 0
    while repeat > 0:
        for i, label in enumerate(labeling_list_type):
            ws.cell(row=2, column=2+i+c, value=label)
        c += 4
        repeat -= 1

    labeling_list_year = ['Spend', '19/20', '20/21', '21/22', '22/23', '23/24', '24/25', '25/26', '26/27', '27/28',
                          '28/29', 'Unprofiled']

    for i, label in enumerate(labeling_list_year):
        ws.cell(row=2+i, column=1, value=label)

    '''process for showing total cost profile. starting with data'''

    row_start = 16
    for x, l in enumerate(total_list):
        for i, total in enumerate(l):
            ws.cell(row=i + row_start, column=x + 2, value=total)

    '''data for graph labeling'''

    for i, quarter in enumerate(labeling_list_quarter):
        ws.cell(row=15, column=i + 2, value=quarter)

    for i, label in enumerate(labeling_list_year):
        ws.cell(row=15+i, column=1, value=label)


    chart = LineChart()
    chart.title = str(name) + ' Cost Profile'
    chart.style = 4
    chart.x_axis.title = 'Financial Year'
    chart.y_axis.title = 'Cost £m'

    '''styling chart'''
    # axis titles
    font = Font(typeface='Calibri')
    size = 1200  # 12 point size
    cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    chart.x_axis.title.tx.rich.p[0].pPr = pp
    chart.y_axis.title.tx.rich.p[0].pPr = pp
    # chart.title.tx.rich.p[0].pPr = pp

    # title
    size_2 = 1400
    cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
    pp_2 = ParagraphProperties(defRPr=cp_2)
    rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
    chart.title.tx.rich.p[0].pPr = pp_2

    '''unprofiled costs not included in the chart'''
    data = Reference(ws, min_col=2, min_row=15, max_col=4, max_row=25)
    chart.add_data(data, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=16, max_row=25)
    chart.set_categories(cats)

    s3 = chart.series[0]
    s3.graphicalProperties.line.solidFill = "cfcfea"  # light blue
    s8 = chart.series[1]
    s8.graphicalProperties.line.solidFill = "5097a4"  # medium blue
    s9 = chart.series[2]
    s9.graphicalProperties.line.solidFill = "0e2f44"  # dark blue'''

    ws.add_chart(chart, "H15")

    '''process for creating income chart'''

    baseline_total_income = calculate_income_totals(name, baseline_fin_data)
    last_q_total_income = calculate_income_totals(name, last_fin_data)
    latest_q_total_income = calculate_income_totals(name, latest_fin_data)

    total_income_list = [baseline_total_income, last_q_total_income, latest_q_total_income]

    if sum(latest_q_total_income) is not 0:
        for x, l in enumerate(total_income_list):
            for i, total in enumerate(l):
                ws.cell(row=i + 32, column=x + 2, value=total)

        '''data for graph labeling'''

        for i, quarter in enumerate(labeling_list_quarter):
            ws.cell(row=32, column=i + 2, value=quarter)

        for i, label in enumerate(labeling_list_year):
            ws.cell(row=32 + i, column=1, value=label)


        '''income graph'''

        chart = LineChart()
        chart.title = str(name) + ' Income Profile'
        chart.style = 4
        chart.x_axis.title = 'Financial Year'
        chart.y_axis.title = 'Cost £m'

        font = Font(typeface='Calibri')
        size = 1200  # 12 point size
        cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
        pp = ParagraphProperties(defRPr=cp)
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        chart.x_axis.title.tx.rich.p[0].pPr = pp
        chart.y_axis.title.tx.rich.p[0].pPr = pp
        # chart.title.tx.rich.p[0].pPr = pp

        # title
        size_2 = 1400
        cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
        pp_2 = ParagraphProperties(defRPr=cp_2)
        rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
        chart.title.tx.rich.p[0].pPr = pp_2

        #unprofiled costs not included in the chart
        data = Reference(ws, min_col=2, min_row=32, max_col=4, max_row=42)
        chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=33, max_row=42)
        chart.set_categories(cats)


        '''
        keeping as colour coding is useful
        s1 = chart.series[0]
        s1.graphicalProperties.line.solidFill = "cfcfea" #light blue
        s2 = chart.series[1]
        s2.graphicalProperties.line.solidFill = "e2f1bb" #light green 
        s3 = chart.series[2]
        s3.graphicalProperties.line.solidFill = "eaba9d" #light red
        s4 = chart.series[3]
        s4.graphicalProperties.line.solidFil = "5097a4" #medium blue
        s5 = chart.series[4]
        s5.graphicalProperties.line.solidFill = "a0db8e" #medium green
        s6 = chart.series[5]
        s6.graphicalProperties.line.solidFill = "b77575" #medium red
        s7 = chart.series[6]
        s7.graphicalProperties.line.solidFil = "0e2f44" #dark blue
        s8 = chart.series[7]
        s8.graphicalProperties.line.solidFill = "29ab87" #dark green
        s9 = chart.series[8]
        s9.graphicalProperties.line.solidFill = "691c1c" #dark red
        '''

        s3 = chart.series[0]
        s3.graphicalProperties.line.solidFill = "e2f1bb"  # light green
        s8 = chart.series[1]
        s8.graphicalProperties.line.solidFill = "a0db8e"  # medium green
        s9 = chart.series[2]
        s9.graphicalProperties.line.solidFill = "29ab87"  # dark green

        ws.add_chart(chart, "H31")

    else:
        pass

    return wb

'''List of financial data keys to capture. This should be amended to years of interest'''
capture_rdel = ['19-20 RDEL Forecast Total', '20-21 RDEL Forecast Total', '21-22 RDEL Forecast Total',
                '22-23 RDEL Forecast Total', '23-24 RDEL Forecast Total', '24-25 RDEL Forecast Total',
                '25-26 RDEL Forecast Total', '26-27 RDEL Forecast Total', '27-28 RDEL Forecast Total',
                '28-29 RDEL Forecast Total', 'Unprofiled RDEL Forecast Total']
capture_cdel = ['19-20 CDEL Forecast Total', '20-21 CDEL Forecast Total', '21-22 CDEL Forecast Total',
                 '22-23 CDEL Forecast Total', '23-24 CDEL Forecast Total', '24-25 CDEL Forecast Total',
                 '25-26 CDEL Forecast Total', '26-27 CDEL Forecast Total', '27-28 CDEL Forecast Total',
                 '28-29 CDEL Forecast Total', 'Unprofiled CDEL Forecast Total']
capture_ng = ['19-20 Forecast Non-Gov', '20-21 Forecast Non-Gov', '21-22 Forecast Non-Gov', '22-23 Forecast Non-Gov',
              '23-24 Forecast Non-Gov', '24-25 Forecast Non-Gov', '25-26 Forecast Non-Gov', '26-27 Forecast Non-Gov',
              '27-28 Forecast Non-Gov', '28-29 Forecast Non-Gov', 'Unprofiled Forecast-Gov']
capture_income =['19-20 Forecast - Income both Revenue and Capital',
                '20-21 Forecast - Income both Revenue and Capital', '21-22 Forecast - Income both Revenue and Capital',
                '22-23 Forecast - Income both Revenue and Capital', '23-24 Forecast - Income both Revenue and Capital',
                '24-25 Forecast - Income both Revenue and Capital', '25-26 Forecast - Income both Revenue and Capital',
                '26-27 Forecast - Income both Revenue and Capital', '27-28 Forecast - Income both Revenue and Capital',
                '28-29 Forecast - Income both Revenue and Capital', 'Unprofiled Forecast Income']

all_data_lists = capture_rdel + capture_cdel + capture_ng + capture_income

'''INSTRUCTIONS FOR RUNNING PROGRAMME'''

'''1) load all master quarter data files here'''
q1_1920 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2019_wip_'
                                   '(18_7_19).xlsx')
q4_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2018.xlsx')
q3_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2018.xlsx')
q2_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_2_2018.xlsx')
q1_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2018.xlsx')
q4_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2017.xlsx')
q3_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2017.xlsx')
q2_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_2_2017.xlsx')
q1_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2017.xlsx')
q4_1617 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2016.xlsx')
q3_1617 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2016.xlsx')

'''2) Include in the below list, as the variable names, those quarters to include in analysis
NOTE - the comparison of financial totals us the 'bespoke' list consistent cost reporting (in nominal figures) was
 only achieved in q1_1819, so it would be incorrect to compare figures beyond this'''
list_of_dicts_all = [q1_1920 ,q4_1819, q3_1819, q2_1819, q1_1819, q4_1718, q3_1718, q2_1718, q1_1718, q4_1617, q3_1617]
list_of_dicts_bespoke = [q1_1920, q4_1819, q3_1819, q2_1819, q1_1819]

'''3) project name list options - this is where the group of interest is specified '''

'''option 1 - all '''
proj_names_all = list(q1_1920.keys())

'''option 2 - a group'''
#TODO write function for filtering list of project names based on group
proj_names_group = ['East Midlands Franchise', 'Rail Franchising Programme', 'West Coast Partnership Franchise']

'''option 3 - bespoke list of projects'''
proj_names_bespoke = ['']

'''4) Enter at the bottom of this function the file path to where outputs should be saved.'''

def run_financials_single(proj_list, q_masters_dict_list):
    baseline_bc = bc_ref_stages(proj_list, q_masters_dict_list)
    q_masters_list = get_master_baseline_dict(proj_list, q_masters_dict_list, baseline_bc)
    latest_financial_data = financial_dict(proj_names_all, q_masters_dict_list, q_masters_list, all_data_lists, 0)
    last_financial_data = financial_dict(proj_names_all, q_masters_dict_list, q_masters_list, all_data_lists, 1)
    baseline_financial_data = financial_dict(proj_names_all, q_masters_dict_list, q_masters_list, all_data_lists, 2)
    for proj_name in proj_list:
        wb = place_in_excel(proj_name, latest_financial_data, last_financial_data, baseline_financial_data)
        wb.save('C:\\Users\\Standalone\\general\\Q1_1920_{}_financial_profile.xlsx'.format(proj_name))


'''5) run the programme placing in the relevant variables. 
the function is structured as run_financials_single(proj_list, q_masters_dict_list)
i) proj_list = list of project names
ii) q_master_dict_list = list of master dictionaries to be passed into the programme'''

run_financials_single(proj_names_all, list_of_dicts_bespoke)