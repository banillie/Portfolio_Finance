'''programme for return financial profile for portfolio'''

from bcompiler.utils import project_data_from_master
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font


'''List of financial data keys to capture data values for analysis'''
capture_rdel = ['18-19 RDEL Forecast Total', '19-20 RDEL Forecast Total',
                 '20-21 RDEL Forecast Total','21-22 RDEL Forecast Total','22-23 RDEL Forecast Total',
                 '23-24 RDEL Forecast Total','24-25 RDEL Forecast Total','25-26 RDEL Forecast Total',
                 '26-27 RDEL Forecast Total','27-28 RDEL Forecast Total','Unprofiled RDEL Forecast Total']


capture_cdel = ['18-19 CDEL Forecast Total','19-20 CDEL Forecast Total',
                '20-21 CDEL Forecast Total','21-22 CDEL Forecast Total',
                 '22-23 CDEL Forecast Total','23-24 CDEL Forecast Total','24-25 CDEL Forecast Total',
                 '25-26 CDEL Forecast Total','26-27 CDEL Forecast Total','27-28 CDEL Forecast Total',
                 'Unprofiled CDEL Forecast Total']

capture_ng = ['18-19 Forecast Non-Gov','19-20 Forecast Non-Gov','20-21 Forecast Non-Gov','21-22 Forecast Non-Gov',
                 '22-23 Forecast Non-Gov','23-24 Forecast Non-Gov','24-25 Forecast Non-Gov',
                 '25-26 Forecast Non-Gov','26-27 Forecast Non-Gov',
                 '27-28 Forecast Non-Gov','Unprofiled Forecast-Gov']

capture_income =['18-19 Forecast - Income both Revenue and Capital', '19-20 Forecast - Income both Revenue and Capital',
                '20-21 Forecast - Income both Revenue and Capital', '21-22 Forecast - Income both Revenue and Capital',
                '22-23 Forecast - Income both Revenue and Capital', '23-24 Forecast - Income both Revenue and Capital',
                '24-25 Forecast - Income both Revenue and Capital', '25-26 Forecast - Income both Revenue and Capital',
                '26-27 Forecast - Income both Revenue and Capital', '27-28 Forecast - Income both Revenue and Capital',
                'Unprofiled Forecast Income']

all_data_lists = capture_rdel + capture_cdel + capture_ng + capture_income

'''function that makes list of tuples containing financial key:valueinfo for each project. cells_to_capture lists 
of interest are stored below'''
def financial_info(dictionary, cells_to_capture):
    output_dicitonary = {}
    for name in dictionary.keys():
        output_list = []
        for item in dictionary[name]:
            if item in cells_to_capture:
                if dictionary[name][item] is None:
                    a = item
                    b = 0
                    c = (a, b)
                    output_list.append(c)
                else:
                    a = item
                    b = dictionary[name][item]
                    c = (a, b)
                    output_list.append(c)

        output_dicitonary[name] = output_list

    return output_dicitonary

'''function that calculates year totals. Returns a list'''
def year_totals(cells_to_capture, fin_dictionary):
    totals = []
    for i in range(0, len(cells_to_capture)):
        key = cells_to_capture[i]
        thesum = 0
        for name in fin_dictionary:
            if name in remove_from_totals:
                pass
            else:
                try:
                    thesum = thesum + fin_dictionary[name][key]
                except TypeError:
                    pass
        totals.append(thesum)

    return totals



def place_in_excel(fin_dictionary, total, cells_to_capture):
    '''create a workbook'''
    wb = Workbook()
    ws = wb.active

    '''place project information into spreadsheet
    using the different dictionaries that have been created'''

    for i, name in enumerate(fin_dictionary.keys()):
        '''lists project names in row one'''
        ws.cell(row=1, column=i + 2).value = name

        '''iterates through financial dictionary - placing financial data in ws'''
        for x in range(0, len(fin_dictionary[name])):
            ws.cell(row=x+2, column=i+2).value = fin_dictionary[name][x][1]

    '''places totals in final column'''
    for i, values in enumerate(total):
        ws.cell(row=i + 2, column=len(fin_dictionary.keys())+2).value = values

    '''places keys into the chart'''
    for i, key in enumerate(cells_to_capture):
        ws.cell(row=i+2, column=1).value = key

    '''data for overall chart'''
    start_row = len(total) + 8
    for x in range(0, int(len(total) / 4)):
        ws.cell(row=start_row, column=2, value=total[x])
        start_row += 1

    start_row = len(total) + 8
    for x in range(int(len(total) / 4), (int(len(total) / 4) * 2)):
        ws.cell(row=start_row, column=3, value=total[x])
        start_row += 1

    start_row = len(total) + 8
    for x in range((int(len(total) / 4) * 2), (int(len(total) / 4) * 3)):
        ws.cell(row=start_row, column=4, value=total[x])
        start_row += 1

    start_row = len(total) + 8
    for x in range((int(len(total) / 4) * 3), int(len(total))):
        ws.cell(row=start_row, column=5, value=total[x])
        start_row += 1

    # this is a hack which is currently hard coded
    start_row = len(total) + 8
    list_of_numbers = [0, 11, 22]  # hard coded part
    total_sum = 0
    for i in range(0, 11):
        for x in list_of_numbers:
            total_sum = total_sum + total[x + i]
            ws.cell(row=start_row, column=6, value=total_sum)
        start_row += 1
        total_sum = 0

    a = len(total) + 7
    ws.cell(row=a, column=2, value='RDEL')
    ws.cell(row=a, column=3, value='CDEL')
    ws.cell(row=a, column=4, value='Non-Gov')
    ws.cell(row=a, column=5, value='Income')
    ws.cell(row=a, column=6, value='Total')

    # ws.cell(row=a+1, column=1, value='17/18')
    ws.cell(row=a + 1, column=1, value='18/19')
    ws.cell(row=a + 2, column=1, value='19/20')
    ws.cell(row=a + 3, column=1, value='20/21')
    ws.cell(row=a + 4, column=1, value='21/22')
    ws.cell(row=a + 5, column=1, value='22/23')
    ws.cell(row=a + 6, column=1, value='23/24')
    ws.cell(row=a + 7, column=1, value='24/25')
    ws.cell(row=a + 8, column=1, value='25/26')
    ws.cell(row=a + 9, column=1, value='26/27')
    ws.cell(row=a + 10, column=1, value='27/28')
    ws.cell(row=a + 11, column=1, value='Unprofiled')

    # TODO fix chart
    chart = LineChart()
    chart.title = 'Portfolio cost profile'
    chart.style = 4
    chart.x_axis.title = 'Financial Year'
    chart.y_axis.title = 'Cost (£m)'
    chart.height = 15  # default is 7.5
    chart.width = 26  # default is 15

    '''styling chart'''
    # axis titles
    font = Font(typeface='Calibri')
    size = 1200  # 12 point size
    cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    chart.x_axis.title.tx.rich.p[0].pPr = pp
    chart.y_axis.title.tx.rich.p[0].pPr = pp

    # title
    size_2 = 1400
    cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
    pp_2 = ParagraphProperties(defRPr=cp_2)
    rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
    chart.title.tx.rich.p[0].pPr = pp_2

    data = Reference(ws, min_col=2, min_row=51, max_col=5, max_row=61)
    cats = Reference(ws, min_col=1, min_row=52, max_row=61)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    s3 = chart.series[0]
    s3.graphicalProperties.line.solidFill = "36708a"  # dark blue
    s8 = chart.series[1]
    s8.graphicalProperties.line.solidFill = "68db8b"  # green
    s9 = chart.series[2]
    s9.graphicalProperties.line.solidFill = "794747"  # dark red
    s9 = chart.series[3]
    s9.graphicalProperties.line.solidFill = "73527f"  # purple

    ws.add_chart(chart, "I52")

    return wb


latest_q = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\master_3_2018.xlsx")
last_q = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx")

remove_from_totals = ['HS2 Phase 2b', 'HS2 Phase1', 'HS2 Phase2a', 'East Midlands Franchise',
                    'South Eastern Rail Franchise Competition', 'West Coast Partnership Franchise',
                     'A66 Full Scheme', 'East Coast Digital Programme',
                      'Manchester North West Quadrant']

finance_data = financial_info(latest_q, all_data_lists)
total_data = year_totals(all_data_lists, latest_q)

output = place_in_excel(finance_data, total_data, all_data_lists)

output.save("C:\\Users\\Standalone\\Will\\Q3_1819_p_fin_profile_like_for_like.xlsx")