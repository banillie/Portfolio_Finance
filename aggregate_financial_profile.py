'''
Programme to create a financial profile for a group of projects i.e. can produce the portfolio profile or a chosen
set of projects profile. It also has the option for comparing like for like projects across projects if necessary

Input documents.
1) Two quarter master data sets

Output documents
2) Excel spreadsheet contain a graph with financial profile

Follow instructions below

'''

from bcompiler.utils import project_data_from_master
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font

''' One of key functions used for calculating which quarter to baseline data from...
Function returns a dictionary structured in the following way project name[('latest quarter info', 'latest bc'), 
('last quarter info', 'last bc'), ('last baseline quarter info', 'last baseline bc'), ('oldest quarter info', 
'oldest bc')] depending on the amount information available in the data. Only the first three key values are returned, 
to ensure consistency (which is helpful later).'''
def bc_ref_stages(proj_list, q_masters_dict_list):

    output_dict = {}

    for name in proj_list:
        #print(name)
        all_list = []      # format [('quarter info': 'bc')] across all masters including project
        bl_list = []        # format ['bc', 'bc'] across all masters. bl_list_2 removes duplicates
        ref_list = []       # format as for all list but only contains the three tuples of interest
        for master in q_masters_dict_list:
            try:
                bc_stage = master[name]['BICC approval point']
                quarter = master[name]['Reporting period (GMPP - Snapshot Date)']
                tuple = (quarter, bc_stage)
                all_list.append(tuple)
            except KeyError:
                pass

        for i in range(0, len(all_list)):
            bl_list.append(all_list[i][1])

        '''below lines of text from stackoverflow. Question, remove duplicates in python list while 
        preserving order'''
        seen = set()
        seen_add = seen.add
        bl_list_2 = [x for x in bl_list if not (x in seen or seen_add(x))]

        ref_list.insert(0, all_list[0])     # puts the latest info into the list first

        try:
            ref_list.insert(1, all_list[1])    # puts that last info into the list
        except IndexError:
            ref_list.insert(1, all_list[0])

        if len(bl_list_2) == 1:                     # puts oldest info into list (as basline if no baseline)
            ref_list.insert(2, all_list[-1])
        else:
            for i in range(0, len(all_list)):      # puts in baseline
                if all_list[i][1] == bl_list[0]:
                    ref_list.insert(2, all_list[i])

        '''there is a hack here i.e. returning only first three in ref_list. There's a bug which I don't fully 
        understand, but this solution is hopefully good enough for now'''
        output_dict[name] = ref_list[0:3]

    return output_dict

'''Another key function used for calcualting which quarter to baseline data from...
Fuction returns a dictionay structured in the following way project_name[n,n,n]. The n (number) values denote where 
the relevant quarter master dictionary is positions in the list of master dictionaries'''
def get_master_baseline_dict(proj_list, q_masters_dict_list, baseline_dict_list):
    output_dict = {}

    for name in proj_list:
        master_q_list = []
        for key in baseline_dict_list[name]:
            for x, master in enumerate(q_masters_dict_list):
                try:
                    quarter = master[name]['Reporting period (GMPP - Snapshot Date)']
                    if quarter == key[0]:
                        master_q_list.append(x)
                except KeyError:
                    pass

        output_dict[name] = master_q_list

    return output_dict


def year_totals(proj_list, proj_list_remove, data_key_list, q_masters_dict_list, q_masters_list, period):
    #TODO convert output into dictionary

    fy_total_list = []

    totals_proj_list = []
    for proj_name in proj_list:
        if proj_name not in proj_list_remove:
            totals_proj_list.append(proj_name)

    for key in data_key_list:
        thesum = 0
        for proj_name in totals_proj_list:
            try:
                if period == 'baseline':
                    to_add = q_masters_dict_list[q_masters_list[proj_name][2]][proj_name][key]
                if period == 'last':
                    to_add = q_masters_dict_list[q_masters_list[proj_name][1]][proj_name][key]
                if period == 'latest':
                    to_add = q_masters_dict_list[q_masters_list[proj_name][0]][proj_name][key]
                thesum = thesum + to_add
            except (TypeError, KeyError):
                pass
        fy_total_list.append(thesum)

    return fy_total_list

def likeforlike():
    '''
    small programme used to filter out projects that are not in both data sets
    :param data_1: most recent quarters data
    :param data_2: less recent quarters data
    :return: a list of projects that are in both data sets
    '''

    one = list(set(q1_1920) - set(q4_1819))
    two = list(set(q4_1819) - set(q1_1920))

    output_list = one + two

    return output_list

def place_in_excel(proj_list, data_key_list, total_data, q_masters_dict_list, q_masters_list, period):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1).value = 'Project'
    for i, proj_name in enumerate(proj_list):
        '''lists project names in row one'''
        ws.cell(row=1, column=i + 2).value = proj_name

        '''iterates through financial dictionary - placing financial data in ws'''
        for x, key in enumerate(data_key_list):
            try:
                if period == 'baseline':
                    ws.cell(row=x+2, column=i+2).value = q_masters_dict_list[q_masters_list[proj_name][2]][proj_name][key]
                if period == 'last':
                    ws.cell(row=x + 2, column=i + 2).value = \
                    q_masters_dict_list[q_masters_list[proj_name][1]][proj_name][key]
                if period == 'latest':
                    ws.cell(row=x + 2, column=i + 2).value = \
                    q_masters_dict_list[q_masters_list[proj_name][0]][proj_name][key]
            except KeyError:
                ws.cell(row=x + 2, column=i + 2).value = 0

    '''places totals in final column. to note because this is a list and not a dictionary as for fin_data there is 
    possibility that data could become unaligned. Whether changing the list of cells_to_capture causes them to become
    unaligned needs to be tested'''
    ws.cell(row=1, column=len(proj_list) + 2).value = 'Total'
    for i, values in enumerate(total_data):
        ws.cell(row=i + 2, column=len(proj_list)+2).value = values

    '''places keys into the chart in the first column'''
    for i, key in enumerate(data_key_list):
        ws.cell(row=i+2, column=1).value = key

    '''information on which projects are not included in totals'''
    ws.cell(row=1, column=len(proj_list) + 4).value = 'Projects that have been removed to avoid double counting'
    for i, project in enumerate(dont_double_count):
        ws.cell(row=i + 2, column=len(proj_list) + 4).value = project

    # ws.cell(row=1, column=len(proj_list)+6).value = 'Projects that have been removed to enable like for like' \
    #                                                       'comparison of totals'
    # for i, project in enumerate(like_for_like_totals):
    #     ws.cell(row=i + 2, column=len(proj_list)+6).value = project

    '''data for overall chart. As above because this data is in a list - possibility of it being unaligned needs 
    testing. not the best way of managing data flow, but working for now'''
    start_row = len(total_data) + 8
    for x in range(0, int(len(total_data) / 4)):
        ws.cell(row=start_row, column=2, value=total_data[x])
        start_row += 1

    start_row = len(total_data) + 8
    for x in range(int(len(total_data) / 4), (int(len(total_data) / 4) * 2)):
        ws.cell(row=start_row, column=3, value=total_data[x])
        start_row += 1

    start_row = len(total_data) + 8
    for x in range((int(len(total_data) / 4) * 2), (int(len(total_data) / 4) * 3)):
        ws.cell(row=start_row, column=4, value=total_data[x])
        start_row += 1

    start_row = len(total_data) + 8
    for x in range((int(len(total_data) / 4) * 3), int(len(total_data))):
        ws.cell(row=start_row, column=5, value=total_data[x])
        start_row += 1

    '''code was essentially a hack'''

    start_row = len(total_data) + 8
    list_of_numbers = [0, len(capture_rdel), len(capture_rdel)*2]
    total_sum = 0
    for i in range(0, len(capture_rdel)):
        for x in list_of_numbers:
            total_sum = total_sum + total_data[x + i]
            ws.cell(row=start_row, column=6, value=total_sum)
        start_row += 1
        total_sum = 0

    a = len(total_data) + 7
    ws.cell(row=a, column=2, value='RDEL')
    ws.cell(row=a, column=3, value='CDEL')
    ws.cell(row=a, column=4, value='Non-Gov')
    ws.cell(row=a, column=5, value='Income')
    ws.cell(row=a, column=6, value='Total')


    # ws.cell(row=a+1, column=1, value='17/18')
    #ws.cell(row=a + 1, column=1, value='18/19')
    ws.cell(row=a + 1, column=1, value='19/20')
    ws.cell(row=a + 2, column=1, value='20/21')
    ws.cell(row=a + 3, column=1, value='21/22')
    ws.cell(row=a + 4, column=1, value='22/23')
    ws.cell(row=a + 5, column=1, value='23/24')
    ws.cell(row=a + 6, column=1, value='24/25')
    ws.cell(row=a + 7, column=1, value='25/26')
    ws.cell(row=a + 8, column=1, value='26/27')
    ws.cell(row=a + 9, column=1, value='27/28')
    ws.cell(row=a + 10, column=1, value='28/29')
    ws.cell(row=a + 11, column=1, value='Unprofiled')

    '''this builds a very basic chart'''
    # TODO fix chart
    chart = LineChart()
    chart.title = 'Portfolio cost profile'
    chart.style = 4
    chart.x_axis.title = 'Financial Year'
    chart.y_axis.title = 'Cost (£m)'
    chart.height = 15  # default is 7.5
    chart.width = 26  # default is 15

    '''styling chart'''
    # axis titles
    font = Font(typeface='Calibri')
    size = 1200  # 12 point size
    cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    chart.x_axis.title.tx.rich.p[0].pPr = pp
    chart.y_axis.title.tx.rich.p[0].pPr = pp

    # title
    size_2 = 1400
    cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
    pp_2 = ParagraphProperties(defRPr=cp_2)
    rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
    chart.title.tx.rich.p[0].pPr = pp_2

    data = Reference(ws, min_col=2, min_row=51, max_col=5, max_row=61)
    cats = Reference(ws, min_col=1, min_row=52, max_row=61)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    s3 = chart.series[0]
    s3.graphicalProperties.line.solidFill = "36708a"  # dark blue
    s8 = chart.series[1]
    s8.graphicalProperties.line.solidFill = "68db8b"  # green
    s9 = chart.series[2]
    s9.graphicalProperties.line.solidFill = "794747"  # dark red
    s9 = chart.series[3]
    s9.graphicalProperties.line.solidFill = "73527f"  # purple

    ws.add_chart(chart, "I52")

    return wb

def run_financials_all(proj_list, proj_list_remove, data_key_list, q_masters_dict_list, period):
    baseline_bc = bc_ref_stages(proj_list, q_masters_dict_list)
    q_masters_list = get_master_baseline_dict(proj_list, q_masters_dict_list, baseline_bc)
    total_data = year_totals(proj_list, proj_list_remove, data_key_list, q_masters_dict_list, q_masters_list, period)
    output = place_in_excel(proj_list, data_key_list, total_data, q_masters_dict_list, q_masters_list, period)

    return output

'''List of financial data keys to capture. This should be amended to years of interest'''
capture_rdel = ['19-20 RDEL Forecast Total', '20-21 RDEL Forecast Total', '21-22 RDEL Forecast Total',
                '22-23 RDEL Forecast Total', '23-24 RDEL Forecast Total', '24-25 RDEL Forecast Total',
                '25-26 RDEL Forecast Total', '26-27 RDEL Forecast Total', '27-28 RDEL Forecast Total',
                '28-29 RDEL Forecast Total', 'Unprofiled RDEL Forecast Total']
capture_cdel = ['19-20 CDEL Forecast Total', '20-21 CDEL Forecast Total', '21-22 CDEL Forecast Total',
                 '22-23 CDEL Forecast Total', '23-24 CDEL Forecast Total', '24-25 CDEL Forecast Total',
                 '25-26 CDEL Forecast Total', '26-27 CDEL Forecast Total', '27-28 CDEL Forecast Total',
                 '28-29 CDEL Forecast Total', 'Unprofiled CDEL Forecast Total']
capture_ng = ['19-20 Forecast Non-Gov', '20-21 Forecast Non-Gov', '21-22 Forecast Non-Gov', '22-23 Forecast Non-Gov',
              '23-24 Forecast Non-Gov', '24-25 Forecast Non-Gov', '25-26 Forecast Non-Gov', '26-27 Forecast Non-Gov',
              '27-28 Forecast Non-Gov', '28-29 Forecast Non-Gov', 'Unprofiled Forecast-Gov']
capture_income =['19-20 Forecast - Income both Revenue and Capital',
                '20-21 Forecast - Income both Revenue and Capital', '21-22 Forecast - Income both Revenue and Capital',
                '22-23 Forecast - Income both Revenue and Capital', '23-24 Forecast - Income both Revenue and Capital',
                '24-25 Forecast - Income both Revenue and Capital', '25-26 Forecast - Income both Revenue and Capital',
                '26-27 Forecast - Income both Revenue and Capital', '27-28 Forecast - Income both Revenue and Capital',
                '28-29 Forecast - Income both Revenue and Capital', 'Unprofiled Forecast Income']

all_data_lists = capture_rdel + capture_cdel + capture_ng + capture_income

'''INSTRUCTION FOR RUNNING PROGRAMME'''

'''1) load all master quarter data files here'''
q1_1920 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2019_wip_'
                                   '(18_7_19).xlsx')
q4_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2018.xlsx')
q3_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2018.xlsx')
q2_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_2_2018.xlsx')
q1_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2018.xlsx')
q4_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2017.xlsx')
q3_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2017.xlsx')
q2_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_2_2017.xlsx')
q1_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2017.xlsx')
q4_1617 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2016.xlsx')
q3_1617 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2016.xlsx')

'''2) Include in the below list, as the variable names, those quarters to include in analysis
NOTE - the comparison of financial totals us the 'bespoke' list consistent cost reporting (in nominal figures) was
 only achieved in q1_1819, so it would be incorrect to compare figures beyond this'''
list_of_dicts_all = [q1_1920 ,q4_1819, q3_1819, q2_1819, q1_1819, q4_1718, q3_1718, q2_1718, q1_1718, q4_1617, q3_1617]
list_of_dicts_bespoke = [q1_1920, q4_1819, q3_1819, q2_1819, q1_1819]

''' 3) Choose appropriate project name list options - this is where the group of interest for the aggregate chart is 
specified. '''
'''option 1 - all '''
proj_names_all = list(q1_1920.keys())
'''option 2 - a group'''
proj_names_group = ['East Midlands Franchise', 'Rail Franchising Programme', 'West Coast Partnership Franchise']
'''option 3 - bespoke list of projects'''
proj_names_bespoke = ['Digital Railway']

'''4) It is important to consider the list of projects that should included within financial totals of each year. There 
are two key things to consider:
i) whether some project cost profiles should be removed to prevent double counting, 
ii) whether you would like to have a like for like comparison between chosen quarters i.e. compare change in cost profile 
for the same set of projects. If you are going to compare project financial against baselines then this is not necessary
as the baseline position will represent the when each project joined the portfolio. So you are in fact comparing like for
like.'''

'''option one - remove projects to stop double counting'''
dont_double_count = ['HS2 Phase 2b', 'HS2 Phase1', 'HS2 Phase2a', 'East Midlands Franchise',
                     'South Eastern Rail Franchise Competition', 'West Coast Partnership Franchise']

'''option two - ensure that only like for like comparision of totals. see point above'''
#like_for_like_totals = dont_double_count + likeforlike()

'''5) enter variables created via options above into function and run programme. 
the function is structured as follows... run_financials_all(proj_list, proj_list_remove, data_key_list, q_masters_dict_list)

1) proj_list = list of projects to include in analysis. 
2) proj_list_remove = list of projects to not be included in total figures. 
3) data_key_list = the list of financial keys. 
4) q_masters_dict_list = the list of master dictionaries to include in analysis. 
5) period = which financial information you want to return. options are 'baseline', 'last', 'latest' '''

output = run_financials_all(proj_names_all, dont_double_count, all_data_lists, list_of_dicts_bespoke, 'baseline')

'''5) specify where to save to output file - excel spreadsheet with graph'''
output.save("C:\\Users\\Standalone\\general\\Q1_1920_financial_profile_baseline_no_hs2.xlsx")