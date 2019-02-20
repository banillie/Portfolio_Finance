'''

programme to create a dashboard which has meta data for each projects funding position.

returns a wb. data placed into a table

'''

#TODO tidy/up consider how can be used in the future

from openpyxl import load_workbook, Workbook
from bcompiler.utils import project_data_from_master

current_Q_dict = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2018.xlsx')


def filter_dictionary(dictionary, category):
    f_dict = {}
    for x in dictionary:
        if dictionary[x][category] != None:
            f_dict[x] = dictionary[x]
    return f_dict


def get_project_names(data):
    project_name_list = []
    for x in data:
        project_name_list.append(x)
    return project_name_list


current_Q_list = get_project_names(current_Q_dict)
# last_Q_list = get_project_names(last_Q_dict)

output_wb = Workbook()
ws = output_wb.active

for x in range(0, len(current_Q_list)):
    ws.cell(row=x + 2, column=2, value=current_Q_list[x])

output_wb.save('C:\\Users\\Standalone\\Will\\throwaway_list.xlsx')

wb = load_workbook('C:\\Users\\Standalone\\Will\\throwaway_list.xlsx')
ws = wb.active

for row_num in range(2, ws.max_row + 1):
    project_name = ws.cell(row=row_num, column=2).value
    print(project_name)
    if project_name in current_Q_dict:
        ws.cell(row=row_num, column=1).value = current_Q_dict[project_name]['DfT Group']
        ws.cell(row=row_num, column=3).value = current_Q_dict[project_name]['Real or Nominal - Baseline']
        ws.cell(row=row_num, column=4).value = current_Q_dict[project_name]['Real or Nominal - Actual/Forecast']
        ws.cell(row=row_num, column=5).value = current_Q_dict[project_name]['Index Year']
        print(current_Q_dict[project_name]['Index Year'])
        ws.cell(row=row_num, column=6).value = current_Q_dict[project_name]['Deflator']
        ws.cell(row=row_num, column=7).value = current_Q_dict[project_name]['Source of Finance']
        ws.cell(row=row_num, column=8).value = current_Q_dict[project_name]['Other Finance type Description']
        ws.cell(row=row_num, column=9).value = current_Q_dict[project_name]['Project cost to closure']
        ws.cell(row=row_num, column=10).value = current_Q_dict[project_name]['RDEL Total Forecast']
        ws.cell(row=row_num, column=11).value = current_Q_dict[project_name]['CDEL Total Forecast']
        ws.cell(row=row_num, column=12).value = current_Q_dict[project_name]['Non-Gov Total Forecast']
        ws.cell(row=row_num, column=13).value = current_Q_dict[project_name]['Total Forecast']
        ws.cell(row=row_num, column=14).value = current_Q_dict[project_name]['Project MM21 Forecast - Actual']

ws.cell(row=1, column=1, value='DfT Group')
ws.cell(row=1, column=2, value='Project')
ws.cell(row=1, column=3, value='Real/Nominal - Baseline')
ws.cell(row=1, column=4, value='Real/Nominal - Forecast')
ws.cell(row=1, column=5, value='Index Year')
ws.cell(row=1, column=6, value='Deflator')
ws.cell(row=1, column=7, value='Source of Finance')
ws.cell(row=1, column=8, value='Other Finance')
ws.cell(row=1, column=9, value='Cost to closure')
ws.cell(row=1, column=10, value='RDEL WLC')
ws.cell(row=1, column=11, value='CDEL WLC')
ws.cell(row=1, column=12, value='AMEY WLC')
ws.cell(row=1, column=13, value='Total WLC')
ws.cell(row=1, column=14, value='Project End date')
ws.cell(row=1, column=15, value='Year spending profile stops')
ws.cell(row=1, column=16, value='Funding arrangement')

wb.save('C:\\Users\\Standalone\\Will\\Q4_funding_details.xlsx')